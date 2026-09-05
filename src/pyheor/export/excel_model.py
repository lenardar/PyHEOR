"""
Formula-based Excel model export for cross-validation.
======================================================

Creates Excel workbooks that independently compute Markov or PSM models
using **Excel formulas**, enabling cross-validation against Python results.

The Excel file contains:

- **Input section**: Parameters, transition matrix, costs, utilities (values)
- **Calculation section**: State trace, costs, QALYs, discounting (formulas)
- **Summary sheet**: Totals from Excel formulas, Python-computed totals,
  and a Difference row so you can instantly verify they match.

Supports:

- **MarkovModel**: Full formula-based trace, including interval-specific
  transition matrices and state/transition-event costs
- **PSMModel**: Explicit survival inputs feeding formula-based state
  probabilities, costs, QALYs, and discounting

Usage
-----
>>> result = model.run_base_case()
>>> ph.export_excel_model(result, "verification.xlsx")
# or
>>> ph.export_excel_model(model, "verification.xlsx")
"""

from __future__ import annotations

import numpy as np

from openpyxl import Workbook
from openpyxl.utils import get_column_letter as CL
from openpyxl.styles import Font, PatternFill

# -------------------------------------------------------------------------
# Styles
# -------------------------------------------------------------------------
_TITLE_FONT = Font(bold=True, size=14)
_SECTION_FONT = Font(bold=True, size=11, underline="single")
_HEADER_FONT = Font(bold=True)
_NOTE_FONT = Font(italic=True, color="999999")
_INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")  # light yellow
_FMT_COST = '#,##0.00'
_FMT_PROB = '0.000000'
_FMT_ICER = '#,##0'


# =====================================================================
# Public API
# =====================================================================

def export_excel_model(model_or_result, filepath: str, params: dict = None):
    """Export a formula-based Excel model for cross-validation.

    Creates an Excel workbook whose calculation cells use **Excel
    formulas** (SUMPRODUCT, SUM, IF, etc.) that independently replicate
    the Python computation. A Summary sheet shows both Excel and Python
    totals plus their difference.

    Parameters
    ----------
    model_or_result : MarkovModel, PSMModel, BaseResult, or PSMBaseResult
        The model (uses base-case params) or a result object.
    filepath : str
        Output path (.xlsx).
    params : dict, optional
        Override parameter values.  Ignored when a result is passed.

    Raises
    ------
    TypeError
        If the model type is not MarkovModel or PSMModel.

    Examples
    --------
    >>> result = model.run_base_case()
    >>> ph.export_excel_model(result, "verification.xlsx")
    """
    from ..models.markov import MarkovModel
    from ..models.psm import PSMModel
    from ..analysis.results import BaseResult, PSMBaseResult

    if isinstance(model_or_result, (BaseResult, PSMBaseResult)):
        model = model_or_result.model
        params = model_or_result.params
    else:
        model = model_or_result
        if params is None:
            params = {name: p.base for name, p in model.params.items()}

    if isinstance(model, MarkovModel):
        with model._attr_param_override(params):
            _validate_markov_excel_support(model, params)
            py_results = model._simulate_single(params)
            _build_markov_excel(model, filepath, params, py_results)
    elif isinstance(model, PSMModel):
        with model._attr_param_override(params):
            _validate_psm_excel_support(model, params)
            py_results = model._simulate_single(params)
            _build_psm_excel(model, filepath, params, py_results)
    else:
        raise TypeError(
            f"Excel model export supports MarkovModel and PSMModel, "
            f"got {type(model).__name__}. "
            f"DES and MicroSim are individual-based and cannot be "
            f"replicated with Excel formulas."
        )


# =====================================================================
# Markov Model
# =====================================================================

def _build_markov_excel(model, filepath, params, py_results):
    """Build formula-based Excel workbook for a Markov model."""
    wb = Workbook()

    n = model.n_states
    states = model.states
    n_cycles = model.n_cycles
    cl_val = model.cycle_length
    dr_c = model.dr_cost
    dr_q = model.dr_qaly
    hcc_method = model.half_cycle_correction  # str or None
    initial_idx = model.initial_state_idx

    transition_cats = [tc["category"] for tc in model._transition_costs]
    cost_cats = list(dict.fromkeys([*model._costs, *transition_cats]))
    n_cats = len(cost_cats)

    # --- Per-strategy sheets ---
    summary_refs = {}

    for s_idx, strategy in enumerate(model.strategy_names):
        label = model.strategy_labels[strategy]
        sname = _safe_sheet(f"Calc_{label}")
        matrices = [
            model._get_transition_matrix(strategy, params, interval)
            for interval in range(n_cycles)
        ]
        is_time_homogeneous = all(
            np.allclose(matrices[0], matrix, atol=1e-10, rtol=0)
            for matrix in matrices[1:]
        )

        if s_idx == 0:
            ws = wb.active
            ws.title = sname
        else:
            ws = wb.create_sheet(title=sname)

        # ============================================
        # INPUT SECTION (yellow-filled cells)
        # ============================================
        r = 1
        ws.cell(r, 1, f"PyHEOR Markov 验证模型 — {label}").font = _TITLE_FONT
        r += 2

        # -- Settings --
        ws.cell(r, 1, "模型设置").font = _SECTION_FONT
        r += 1
        ROW_DR_C = _write_setting(ws, r, "Discount Rate (Costs)", dr_c); r += 1
        ROW_DR_Q = _write_setting(ws, r, "Discount Rate (QALYs)", dr_q); r += 1
        ROW_CL   = _write_setting(ws, r, "Cycle Length (years)", cl_val); r += 1
        _write_setting(ws, r, "Discount Convention", model.discount_convention); r += 1
        _write_setting(ws, r, "N Cycles", n_cycles); r += 1
        _write_setting(ws, r, "Half-cycle Correction", hcc_method or "No"); r += 1
        _write_setting(ws, r, "Initial State", states[initial_idx]); r += 2

        # -- Transition matrices (transposed: rows=to, cols=from) --
        matrix_label = (
            "转移概率矩阵 (P^T: 行=目标态, 列=来源态)"
            if is_time_homogeneous
            else "各区间转移概率矩阵 (P_t^T: 行=目标态, 列=来源态)"
        )
        ws.cell(r, 1, matrix_label).font = _SECTION_FONT
        r += 1
        matrices_to_write = matrices[:1] if is_time_homogeneous else matrices
        matrix_cells = []
        for interval, matrix in enumerate(matrices_to_write):
            if not is_time_homogeneous:
                ws.cell(r, 1, f"Interval {interval}").font = _HEADER_FONT
                r += 1
            ws.cell(r, 1, "To \\ From")
            for j in range(n):
                ws.cell(r, 2 + j, states[j]).font = _HEADER_FONT
            r += 1
            first_matrix_row = r
            for to_idx in range(n):
                ws.cell(r, 1, states[to_idx])
                for from_idx in range(n):
                    cell = ws.cell(r, 2 + from_idx, matrix[from_idx, to_idx])
                    cell.fill = _INPUT_FILL
                    cell.number_format = _FMT_PROB
                r += 1
            matrix_range = (
                f"{CL(2)}{first_matrix_row}:"
                f"{CL(1 + n)}{first_matrix_row + n - 1}"
            )
            ws.cell(r, 1, "Validation")
            sum_checks = ",".join(
                f"ABS(SUM({CL(2 + from_idx)}{first_matrix_row}:"
                f"{CL(2 + from_idx)}{first_matrix_row + n - 1})-1)<1E-8"
                for from_idx in range(n)
            )
            ws.cell(
                r, 2,
                f'=IF(AND(MIN({matrix_range})>=0,MAX({matrix_range})<=1,'
                f'{sum_checks}),"OK","ERROR")',
            )
            matrix_cells.append([
                [
                    f"${CL(2 + from_idx)}${first_matrix_row + to_idx}"
                    for from_idx in range(n)
                ]
                for to_idx in range(n)
            ])
            r += 1
        if is_time_homogeneous:
            matrix_cells *= n_cycles

        # -- Transition-event costs --
        transition_inputs = []
        if model._transition_costs:
            ws.cell(r, 1, "转移事件费用 (期末发生；列为事件后间隔)").font = _SECTION_FONT
            r += 1
            ws.cell(r, 1, "Category / Transition").font = _HEADER_FONT
            max_schedule = max(
                len(model._get_tc_schedule(tc_def, strategy, params))
                for tc_def in model._transition_costs
            )
            for offset in range(max_schedule):
                ws.cell(r, 2 + offset, f"Offset {offset}").font = _HEADER_FONT
            r += 1
            for tc_def in model._transition_costs:
                schedule = model._get_tc_schedule(tc_def, strategy, params)
                ws.cell(
                    r, 1,
                    f"{tc_def['category']}: "
                    f"{tc_def['from_state']} → {tc_def['to_state']}",
                )
                refs = []
                for offset, amount in enumerate(schedule):
                    cell = ws.cell(r, 2 + offset, amount)
                    cell.fill = _INPUT_FILL
                    cell.number_format = _FMT_COST
                    refs.append(f"${CL(2 + offset)}${r}")
                transition_inputs.append({
                    "category": tc_def["category"],
                    "from_idx": tc_def["from_idx"],
                    "to_idx": tc_def["to_idx"],
                    "schedule_refs": refs,
                })
                r += 1
            r += 1

        # -- State Costs --
        ws.cell(r, 1, "状态费用 (年度费率)").font = _SECTION_FONT
        r += 1
        for j in range(n):
            ws.cell(r, 2 + j, states[j]).font = _HEADER_FONT
        r += 1

        cost_input_rows = {}
        for cat in model._costs:
            vec = model._resolve_state_values(model._costs[cat].values, strategy, params, 0)
            ws.cell(r, 1, cat)
            for j in range(n):
                c = ws.cell(r, 2 + j, vec[j])
                c.fill = _INPUT_FILL
                c.number_format = _FMT_COST
            cost_input_rows[cat] = r
            r += 1
        if not model._costs:
            ws.cell(r, 1, "(无)")
            r += 1
        r += 1

        # -- Utility Weights --
        ws.cell(r, 1, "效用权重").font = _SECTION_FONT
        r += 1
        for j in range(n):
            ws.cell(r, 2 + j, states[j]).font = _HEADER_FONT
        r += 1
        u_vec = model._get_utilities(strategy, params, 0)
        ws.cell(r, 1, "Utility")
        for j in range(n):
            c = ws.cell(r, 2 + j, u_vec[j])
            c.fill = _INPUT_FILL
        UTIL_ROW = r
        r += 1

        # -- Alive mask --
        ws.cell(r, 1, "Alive")
        alive = [1.0 if i in model._alive_states else 0.0 for i in range(n)]
        for j in range(n):
            c = ws.cell(r, 2 + j, alive[j])
            c.fill = _INPUT_FILL
        ALIVE_ROW = r
        r += 2

        # ============================================
        # CALCULATION TABLE (green-ish formulas)
        # ============================================
        ws.cell(r, 1, "计算区 (以下均为 Excel 公式)").font = _SECTION_FONT
        r += 1

        HDR = r  # header row
        D0 = HDR + 1  # first data row (cycle 0)

        # -- Column mapping --
        COL_CYC = 1
        COL_TIM = 2
        tc = 3                                     # trace col start
        te = tc + n - 1                            # trace col end
        COL_RS = te + 1                            # row sum
        COL_VALID = COL_RS + 1                     # trace validation
        COL_DFC = COL_VALID + 1                    # DF cost
        COL_DFQ = COL_DFC + 1                      # DF qaly
        COL_HCC = COL_DFQ + 1

        craw = {}; cdisc = {}
        cc = COL_HCC + 1
        for cat in cost_cats:
            craw[cat] = cc; cdisc[cat] = cc + 1
            cc += 2

        COL_TC = cc;  cc += 1         # total discounted cost
        COL_QR = cc;  cc += 1         # QALY raw
        COL_QD = cc;  cc += 1         # QALY discounted
        COL_LR = cc;  cc += 1         # LY raw
        COL_LD = cc                   # LY discounted

        # -- Headers --
        hdrs = [
            (COL_CYC, "Cycle"), (COL_TIM, "Time(yr)"),
        ]
        for j in range(n):
            hdrs.append((tc + j, f"P({states[j]})"))
        hdrs += [
            (COL_RS, "RowSum"), (COL_VALID, "Validation"),
            (COL_DFC, "DF(cost)"),
            (COL_DFQ, "DF(qaly)"), (COL_HCC, "HCC"),
        ]
        for cat in cost_cats:
            hdrs += [(craw[cat], f"{cat}(raw)"), (cdisc[cat], f"{cat}(disc)")]
        hdrs += [
            (COL_TC, "TotalCost(disc)"),
            (COL_QR, "QALY(raw)"), (COL_QD, "QALY(disc)"),
            (COL_LR, "LY(raw)"), (COL_LD, "LY(disc)"),
        ]
        for col, txt in hdrs:
            ws.cell(HDR, col, txt).font = _HEADER_FONT

        # Refs that stay constant across rows
        drc = f"$B${ROW_DR_C}"
        drq = f"$B${ROW_DR_Q}"
        cl_ref = f"$B${ROW_CL}"
        # cost vector refs  {cat: "$B$rr:${CL(1+n)}$rr"}
        cvr = {
            cat: f"${CL(2)}${cost_input_rows[cat]}:${CL(1+n)}${cost_input_rows[cat]}"
            for cat in model._costs
        }
        util_r = f"${CL(2)}${UTIL_ROW}:${CL(1+n)}${UTIL_ROW}"
        alive_r = f"${CL(2)}${ALIVE_ROW}:${CL(1+n)}${ALIVE_ROW}"

        # P_t^T row references, one set for each model interval.
        pt_rows = [
            [f"{cells[to_idx][0]}:{cells[to_idx][-1]}" for to_idx in range(n)]
            for cells in matrix_cells
        ]

        # -- Data Rows --
        for t in range(n_cycles + 1):
            rr = D0 + t  # current row

            # Cycle & Time
            ws.cell(rr, COL_CYC, t)
            ws.cell(rr, COL_TIM, f"={CL(COL_CYC)}{rr}*{cl_ref}")

            # Trace ---
            tr_range = f"{CL(tc)}{rr}:{CL(te)}{rr}"  # same-row trace (for costs)
            if t == 0:
                for j in range(n):
                    ws.cell(rr, tc + j, 1.0 if j == initial_idx else 0.0)
            else:
                prev_tr = f"{CL(tc)}{rr - 1}:{CL(te)}{rr - 1}"
                for j in range(n):
                    ws.cell(rr, tc + j,
                            f"=SUMPRODUCT({prev_tr},{pt_rows[t - 1][j]})")

            # Row Sum
            ws.cell(rr, COL_RS, f"=SUM({CL(tc)}{rr}:{CL(te)}{rr})")
            ws.cell(
                rr, COL_VALID,
                f'=IF(AND(MIN({CL(tc)}{rr}:{CL(te)}{rr})>=-1E-10,'
                f'ABS({CL(COL_RS)}{rr}-1)<1E-8),"OK","ERROR")',
            )

            # The final row is an observation point only. It has no rewards.
            if t == n_cycles:
                continue

            # Flow rewards occur at the interval midpoint.
            time_c = f"({CL(COL_TIM)}{rr}+{cl_ref}/2)"
            ws.cell(
                rr, COL_DFC,
                _discount_formula(time_c, drc, model.discount_convention),
            )
            ws.cell(
                rr, COL_DFQ,
                _discount_formula(time_c, drq, model.discount_convention),
            )
            ws.cell(
                rr, COL_HCC,
                "Average endpoints" if hcc_method == "trapezoidal" else "Start",
            )

            df_c_cell = f"{CL(COL_DFC)}{rr}"
            df_q_cell = f"{CL(COL_DFQ)}{rr}"
            next_tr_range = f"{CL(tc)}{rr + 1}:{CL(te)}{rr + 1}"
            if hcc_method == "trapezoidal":
                occupancy = f"({tr_range}+{next_tr_range})/2"
            else:
                occupancy = tr_range

            # --- Costs ---
            for cat in cost_cats:
                state_raw = "0"
                state_disc = "0"
                if cat in model._costs:
                    cdef = model._costs[cat]
                    base = f"SUMPRODUCT({occupancy},{cvr[cat]})"
                    inner = f"{base}*{cl_ref}"

                    if cdef.first_cycle_only:
                        state_raw = f"IF({CL(COL_CYC)}{rr}=0,{inner},0)"
                    elif cdef.apply_cycles is not None:
                        checks = ",".join(
                            f"{CL(COL_CYC)}{rr}={cycle}"
                            for cycle in cdef.apply_cycles
                        )
                        condition = f"OR({checks})" if checks else "FALSE"
                        state_raw = f"IF({condition},{inner},0)"
                    else:
                        state_raw = inner

                    if cdef.method == "starting":
                        starting = f"SUMPRODUCT({tr_range},{cvr[cat]})"
                        state_raw = f"IF({CL(COL_CYC)}{rr}=0,{starting},0)"
                        state_disc = state_raw
                    else:
                        state_disc = f"({state_raw})*{df_c_cell}"

                event_terms = []
                for event in transition_inputs:
                    if event["category"] != cat:
                        continue
                    for offset, amount_ref in enumerate(event["schedule_refs"]):
                        source_interval = t - offset
                        if source_interval < 0:
                            continue
                        probability = matrix_cells[source_interval][
                            event["to_idx"]
                        ][event["from_idx"]]
                        source_row = D0 + source_interval
                        source_trace = f"{CL(tc + event['from_idx'])}{source_row}"
                        event_terms.append(
                            f"{source_trace}*{probability}*{amount_ref}"
                        )
                event_raw = "+".join(event_terms) if event_terms else "0"
                event_time = f"({CL(COL_TIM)}{rr}+{cl_ref})"
                event_df = _discount_formula(
                    event_time, drc, model.discount_convention,
                )[1:]

                ws.cell(rr, craw[cat], f"={state_raw}+({event_raw})")
                ws.cell(rr, craw[cat]).number_format = _FMT_COST
                ws.cell(
                    rr, cdisc[cat],
                    f"={state_disc}+({event_raw})*({event_df})",
                )
                ws.cell(rr, cdisc[cat]).number_format = _FMT_COST

            # Total discounted cost
            if n_cats:
                expr = "+".join(f"{CL(cdisc[c])}{rr}" for c in cost_cats)
                ws.cell(rr, COL_TC, f"={expr}")
            else:
                ws.cell(rr, COL_TC, 0)
            ws.cell(rr, COL_TC).number_format = _FMT_COST

            # --- QALYs ---
            ws.cell(rr, COL_QR,
                    f"=SUMPRODUCT({occupancy},{util_r})*{cl_ref}")
            ws.cell(rr, COL_QD,
                    f"={CL(COL_QR)}{rr}*{df_q_cell}")
            ws.cell(rr, COL_QD).number_format = _FMT_PROB

            # --- LYs ---
            ws.cell(rr, COL_LR,
                    f"=SUMPRODUCT({occupancy},{alive_r})*{cl_ref}")
            ws.cell(rr, COL_LD,
                    f"={CL(COL_LR)}{rr}*{df_q_cell}")

        # -- Totals Row --
        TR = D0 + n_cycles + 1
        DL = D0 + n_cycles - 1  # last interval row; final trace row has no rewards
        ws.cell(TR, COL_CYC, "TOTAL").font = _HEADER_FONT

        sum_cols = (
            [cdisc[c] for c in cost_cats]
            + [COL_TC, COL_QD, COL_LD]
        )
        for sc in sum_cols:
            ws.cell(TR, sc,
                    f"=SUM({CL(sc)}{D0}:{CL(sc)}{DL})")
            ws.cell(TR, sc).font = _HEADER_FONT
            ws.cell(TR, sc).number_format = _FMT_COST

        # Store refs for Summary
        summary_refs[strategy] = {
            'sheet': sname,
            'cost': f"'{sname}'!{CL(COL_TC)}{TR}",
            'qaly': f"'{sname}'!{CL(COL_QD)}{TR}",
            'ly':   f"'{sname}'!{CL(COL_LD)}{TR}",
        }

        # Freeze & widths
        ws.freeze_panes = ws.cell(HDR + 1, tc)
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 10

    # ============================================
    # SUMMARY SHEET
    # ============================================
    _build_summary_sheet(wb, model, summary_refs, py_results)

    _enable_excel_recalculation(wb)
    wb.save(filepath)
    print(f"✅ Excel 验证模型已导出: {filepath}")


# =====================================================================
# PSM Model
# =====================================================================

def _build_psm_excel(model, filepath, params, py_results):
    """Build formula-based Excel workbook for a PSM model."""
    wb = Workbook()

    n = model.n_states
    states = model.states
    n_ep = model.n_endpoints
    endpoints = model.survival_endpoints
    n_cycles = model.n_cycles
    cl_val = model.cycle_length
    dr_c = model.dr_cost
    dr_q = model.dr_qaly
    hcc_method = model.half_cycle_correction  # str or None

    cost_cats = list(model._costs.keys())
    n_cats = len(cost_cats)

    summary_refs = {}

    for s_idx, strategy in enumerate(model.strategy_names):
        label = model.strategy_labels[strategy]
        sname = _safe_sheet(f"Calc_{label}")

        if s_idx == 0:
            ws = wb.active
            ws.title = sname
        else:
            ws = wb.create_sheet(title=sname)

        # ============================================
        # INPUT SECTION
        # ============================================
        r = 1
        ws.cell(r, 1, f"PyHEOR PSM 验证模型 — {label}").font = _TITLE_FONT
        r += 2

        ws.cell(r, 1, "模型设置").font = _SECTION_FONT; r += 1
        ROW_DR_C = _write_setting(ws, r, "Discount Rate (Costs)", dr_c); r += 1
        ROW_DR_Q = _write_setting(ws, r, "Discount Rate (QALYs)", dr_q); r += 1
        ROW_CL   = _write_setting(ws, r, "Cycle Length (years)", cl_val); r += 1
        _write_setting(ws, r, "Discount Convention", model.discount_convention); r += 1
        _write_setting(ws, r, "N Cycles", n_cycles); r += 1
        _write_setting(ws, r, "Half-cycle Correction", hcc_method or "No"); r += 2

        # Survival parameters are Excel inputs when the distribution has a
        # compact, standard spreadsheet formula. Other curve types remain
        # explicit external survival inputs in the calculation table.
        ws.cell(r, 1, "生存曲线参数").font = _SECTION_FONT; r += 1
        curve_specs = {}
        for endpoint in endpoints:
            curve = model._resolve_curve(strategy, endpoint, params)
            spec, r = _write_survival_curve_inputs(
                ws, r, endpoint, curve,
            )
            curve_specs[endpoint] = spec
        r += 1

        # State Costs
        ws.cell(r, 1, "状态费用 (年度费率)").font = _SECTION_FONT; r += 1
        for j in range(n):
            ws.cell(r, 2 + j, states[j]).font = _HEADER_FONT
        r += 1
        cost_input_rows = {}
        for cat in cost_cats:
            vec = model._resolve_state_values(model._costs[cat].values, strategy, params, 0)
            ws.cell(r, 1, cat)
            for j in range(n):
                c = ws.cell(r, 2 + j, vec[j])
                c.fill = _INPUT_FILL
                c.number_format = _FMT_COST
            cost_input_rows[cat] = r
            r += 1
        if not cost_cats:
            ws.cell(r, 1, "(无)"); r += 1
        r += 1

        # Utility
        ws.cell(r, 1, "效用权重").font = _SECTION_FONT; r += 1
        for j in range(n):
            ws.cell(r, 2 + j, states[j]).font = _HEADER_FONT
        r += 1
        u_vec = model._get_utilities(strategy, params, 0)
        ws.cell(r, 1, "Utility")
        for j in range(n):
            c = ws.cell(r, 2 + j, u_vec[j])
            c.fill = _INPUT_FILL
        UTIL_ROW = r; r += 1

        ws.cell(r, 1, "Alive")
        alive = [1.0 if i in model._alive_states else 0.0 for i in range(n)]
        for j in range(n):
            c = ws.cell(r, 2 + j, alive[j])
            c.fill = _INPUT_FILL
        ALIVE_ROW = r; r += 2

        # ============================================
        # CALCULATION TABLE
        # ============================================
        ws.cell(
            r, 1,
            "计算区 (标准参数曲线为公式；其余生存率为明示外部输入)",
        ).font = _SECTION_FONT
        r += 1

        HDR = r
        D0 = HDR + 1

        # Column layout
        COL_CYC = 1
        COL_TIM = 2
        # Survival curves (values)
        surv_start = 3
        surv_end = surv_start + n_ep - 1
        # State probabilities (formulas)
        sp_start = surv_end + 1
        sp_end = sp_start + n - 1
        COL_RS = sp_end + 1
        COL_VALID = COL_RS + 1
        COL_DFC = COL_VALID + 1
        COL_DFQ = COL_DFC + 1
        COL_HCC = COL_DFQ + 1

        craw = {}; cdisc = {}
        cc = COL_HCC + 1
        for cat in cost_cats:
            craw[cat] = cc; cdisc[cat] = cc + 1; cc += 2

        COL_TC = cc; cc += 1
        COL_QR = cc; cc += 1
        COL_QD = cc; cc += 1
        COL_LR = cc; cc += 1
        COL_LD = cc

        # Headers
        headers = [(COL_CYC, "Cycle"), (COL_TIM, "Time(yr)")]
        for j, ep in enumerate(endpoints):
            headers.append((surv_start + j, f"S({ep})"))
        for j, st in enumerate(states):
            headers.append((sp_start + j, f"P({st})"))
        headers += [
            (COL_RS, "RowSum"), (COL_VALID, "Validation"),
            (COL_DFC, "DF(cost)"),
            (COL_DFQ, "DF(qaly)"), (COL_HCC, "HCC"),
        ]
        for cat in cost_cats:
            headers += [(craw[cat], f"{cat}(raw)"), (cdisc[cat], f"{cat}(disc)")]
        headers += [
            (COL_TC, "TotalCost(disc)"),
            (COL_QR, "QALY(raw)"), (COL_QD, "QALY(disc)"),
            (COL_LR, "LY(raw)"), (COL_LD, "LY(disc)"),
        ]
        for col, txt in headers:
            ws.cell(HDR, col, txt).font = _HEADER_FONT

        # Constant refs
        drc = f"$B${ROW_DR_C}"
        drq = f"$B${ROW_DR_Q}"
        cl_ref = f"$B${ROW_CL}"
        cvr = {cat: f"${CL(2)}${cost_input_rows[cat]}:${CL(1+n)}${cost_input_rows[cat]}"
               for cat in cost_cats}
        util_r = f"${CL(2)}${UTIL_ROW}:${CL(1+n)}${UTIL_ROW}"
        alive_r = f"${CL(2)}${ALIVE_ROW}:${CL(1+n)}${ALIVE_ROW}"

        # Get survival values from Python
        surv_data = py_results[strategy]['survival_curves']  # {ep: array}

        # -- Data Rows --
        for t in range(n_cycles + 1):
            rr = D0 + t

            ws.cell(rr, COL_CYC, t)
            ws.cell(rr, COL_TIM, f"={CL(COL_CYC)}{rr}*{cl_ref}")

            # Survival curve formulas where supported; otherwise values are
            # deliberately exposed as editable external inputs.
            for j, ep in enumerate(endpoints):
                spec = curve_specs[ep]
                if spec is None:
                    cell = ws.cell(
                        rr, surv_start + j, float(surv_data[ep][t]),
                    )
                    cell.fill = _INPUT_FILL
                else:
                    time_ref = f"{CL(COL_TIM)}{rr}"
                    cell = ws.cell(
                        rr, surv_start + j,
                        _survival_formula(spec, time_ref),
                    )
                cell.number_format = _FMT_PROB

            # State probability FORMULAS
            # state[0] = S(endpoint_0)
            ws.cell(rr, sp_start,
                    f"={CL(surv_start)}{rr}")

            # Invalid curve ordering is rejected by Python before export, so
            # Excel shows the actual subtraction rather than silently clipping.
            for k in range(1, n_ep):
                ws.cell(rr, sp_start + k,
                        f"={CL(surv_start + k)}{rr}-{CL(surv_start + k - 1)}{rr}")

            # state[-1] = 1 - S(last_endpoint)
            ws.cell(rr, sp_start + n - 1,
                    f"=1-{CL(surv_end)}{rr}")

            # Row Sum
            ws.cell(rr, COL_RS,
                    f"=SUM({CL(sp_start)}{rr}:{CL(sp_end)}{rr})")
            monotonic = ""
            if t > 0:
                checks = ",".join(
                    f"{CL(surv_start + j)}{rr}<="
                    f"{CL(surv_start + j)}{rr - 1}+1E-10"
                    for j in range(n_ep)
                )
                monotonic = f",{checks}"
            ws.cell(
                rr, COL_VALID,
                f'=IF(AND(MIN({CL(surv_start)}{rr}:{CL(surv_end)}{rr})>=0,'
                f'MAX({CL(surv_start)}{rr}:{CL(surv_end)}{rr})<=1,'
                f'MIN({CL(sp_start)}{rr}:{CL(sp_end)}{rr})>=-1E-10,'
                f'ABS({CL(COL_RS)}{rr}-1)<1E-8{monotonic}),"OK","ERROR")',
            )

            if t == n_cycles:
                continue

            time_c = f"({CL(COL_TIM)}{rr}+{cl_ref}/2)"
            ws.cell(
                rr, COL_DFC,
                _discount_formula(time_c, drc, model.discount_convention),
            )
            ws.cell(
                rr, COL_DFQ,
                _discount_formula(time_c, drq, model.discount_convention),
            )
            ws.cell(
                rr, COL_HCC,
                "Average endpoints" if hcc_method == "trapezoidal" else "Start",
            )

            df_c_cell = f"{CL(COL_DFC)}{rr}"
            df_q_cell = f"{CL(COL_DFQ)}{rr}"
            sp_range = f"{CL(sp_start)}{rr}:{CL(sp_end)}{rr}"
            next_sp_range = f"{CL(sp_start)}{rr + 1}:{CL(sp_end)}{rr + 1}"
            if hcc_method == "trapezoidal":
                occupancy = f"({sp_range}+{next_sp_range})/2"
            else:
                occupancy = sp_range

            # Costs
            for cat in cost_cats:
                cdef = model._costs[cat]
                base = f"SUMPRODUCT({occupancy},{cvr[cat]})"

                if cdef.method == "wlos":
                    inner = f"{base}*{cl_ref}"
                else:
                    inner = base

                if cdef.first_cycle_only:
                    raw_f = f"=IF({CL(COL_CYC)}{rr}=0,{inner},0)"
                elif cdef.apply_cycles is not None:
                    checks = ",".join(
                        f"{CL(COL_CYC)}{rr}={cycle}"
                        for cycle in cdef.apply_cycles
                    )
                    condition = f"OR({checks})" if checks else "FALSE"
                    raw_f = f"=IF({condition},{inner},0)"
                else:
                    raw_f = f"={inner}"

                if cdef.method == "starting":
                    starting = f"SUMPRODUCT({sp_range},{cvr[cat]})"
                    raw_f = (
                        f"=IF({CL(COL_CYC)}{rr}=0,{starting},0)"
                    )

                ws.cell(rr, craw[cat], raw_f)
                ws.cell(rr, craw[cat]).number_format = _FMT_COST
                raw_ref = f"{CL(craw[cat])}{rr}"
                if cdef.method == "starting":
                    ws.cell(rr, cdisc[cat], f"={raw_ref}")
                else:
                    ws.cell(rr, cdisc[cat], f"={raw_ref}*{df_c_cell}")
                ws.cell(rr, cdisc[cat]).number_format = _FMT_COST

            # Total discounted cost
            if n_cats:
                expr = "+".join(f"{CL(cdisc[c])}{rr}" for c in cost_cats)
                ws.cell(rr, COL_TC, f"={expr}")
            else:
                ws.cell(rr, COL_TC, 0)
            ws.cell(rr, COL_TC).number_format = _FMT_COST

            # QALYs
            ws.cell(rr, COL_QR,
                    f"=SUMPRODUCT({occupancy},{util_r})*{cl_ref}")
            ws.cell(rr, COL_QD,
                    f"={CL(COL_QR)}{rr}*{df_q_cell}")
            ws.cell(rr, COL_QD).number_format = _FMT_PROB

            # LYs
            ws.cell(rr, COL_LR,
                    f"=SUMPRODUCT({occupancy},{alive_r})*{cl_ref}")
            ws.cell(rr, COL_LD,
                    f"={CL(COL_LR)}{rr}*{df_q_cell}")

        # Totals
        TR = D0 + n_cycles + 1
        DL = D0 + n_cycles - 1
        ws.cell(TR, COL_CYC, "TOTAL").font = _HEADER_FONT
        for sc in ([cdisc[c] for c in cost_cats]
                   + [COL_TC, COL_QD, COL_LD]):
            ws.cell(TR, sc,
                    f"=SUM({CL(sc)}{D0}:{CL(sc)}{DL})")
            ws.cell(TR, sc).font = _HEADER_FONT
            ws.cell(TR, sc).number_format = _FMT_COST

        summary_refs[strategy] = {
            'sheet': sname,
            'cost': f"'{sname}'!{CL(COL_TC)}{TR}",
            'qaly': f"'{sname}'!{CL(COL_QD)}{TR}",
            'ly':   f"'{sname}'!{CL(COL_LD)}{TR}",
        }

        ws.freeze_panes = ws.cell(HDR + 1, surv_start)
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 10

    # Summary
    _build_summary_sheet(wb, model, summary_refs, py_results)

    _enable_excel_recalculation(wb)
    wb.save(filepath)
    print(f"✅ Excel 验证模型已导出: {filepath}")
    print("   ℹ 标准参数生存曲线使用 Excel 公式；其他曲线的"
          "生存率标为外部输入。")


# =====================================================================
# Summary Sheet (shared by Markov & PSM)
# =====================================================================

def _build_summary_sheet(wb, model, summary_refs, py_results):
    """Build a Summary & Cross-Validation sheet."""
    ws = wb.create_sheet("Summary", 0)  # first sheet
    ws.cell(1, 1, "Summary & Cross-Validation").font = _TITLE_FONT

    r = 3
    # --- Excel formula results ---
    ws.cell(r, 1, "Excel 公式计算结果").font = _SECTION_FONT; r += 1
    for h, c in [("Strategy", 1), ("Total Cost", 2),
                 ("QALYs", 3), ("LYs", 4)]:
        ws.cell(r, c, h).font = _HEADER_FONT
    excel_rows = {}
    for s_idx, strategy in enumerate(model.strategy_names):
        rr = r + 1 + s_idx
        ws.cell(rr, 1, model.strategy_labels[strategy])
        ws.cell(rr, 2, f"={summary_refs[strategy]['cost']}")
        ws.cell(rr, 2).number_format = _FMT_COST
        ws.cell(rr, 3, f"={summary_refs[strategy]['qaly']}")
        ws.cell(rr, 3).number_format = _FMT_PROB
        ws.cell(rr, 4, f"={summary_refs[strategy]['ly']}")
        ws.cell(rr, 4).number_format = _FMT_PROB
        excel_rows[strategy] = rr

    # --- ICER ---
    r = rr + 2
    ws.cell(r, 1, "ICER (Excel 公式)").font = _SECTION_FONT; r += 1
    for h, c in [("Strategy", 1), ("vs", 2), ("Inc. Cost", 3),
                 ("Inc. QALYs", 4), ("ICER / Classification", 5)]:
        ws.cell(r, c, h).font = _HEADER_FONT

    comp = model.strategy_names[0]
    comp_row = excel_rows[comp]
    for strategy in model.strategy_names[1:]:
        r += 1
        ir = excel_rows[strategy]
        ws.cell(r, 1, model.strategy_labels[strategy])
        ws.cell(r, 2, model.strategy_labels[comp])
        ws.cell(r, 3, f"=B{ir}-B{comp_row}")
        ws.cell(r, 3).number_format = _FMT_COST
        ws.cell(r, 4, f"=C{ir}-C{comp_row}")
        ws.cell(r, 4).number_format = _FMT_PROB
        ws.cell(
            r, 5,
            f'=IF(ABS(D{r})<0.0000000001,'
            f'IF(C{r}>0.0000000001,"Dominated",'
            f'IF(C{r}<-0.0000000001,"Dominant","No difference")),'
            f'IF(AND(D{r}>0,C{r}<=0),"Dominant",'
            f'IF(AND(D{r}<0,C{r}>=0),"Dominated",C{r}/D{r})))'
        )
        ws.cell(r, 5).number_format = _FMT_ICER

    # --- Python results ---
    r += 2
    ws.cell(r, 1, "Python 计算结果 (对照)").font = _SECTION_FONT; r += 1
    for h, c in [("Strategy", 1), ("Total Cost", 2),
                 ("QALYs", 3), ("LYs", 4)]:
        ws.cell(r, c, h).font = _HEADER_FONT
    py_rows = {}
    for s_idx, strategy in enumerate(model.strategy_names):
        rr = r + 1 + s_idx
        pr = py_results[strategy]
        ws.cell(rr, 1, model.strategy_labels[strategy])
        ws.cell(rr, 2, sum(pr['total_costs'].values()))
        ws.cell(rr, 2).number_format = _FMT_COST
        ws.cell(rr, 3, pr['total_qalys'])
        ws.cell(rr, 3).number_format = _FMT_PROB
        ws.cell(rr, 4, pr['total_lys'])
        ws.cell(rr, 4).number_format = _FMT_PROB
        py_rows[strategy] = rr

    # --- Difference ---
    r = rr + 2
    ws.cell(r, 1, "差异 (Excel − Python)").font = _SECTION_FONT; r += 1
    for h, c in [("Strategy", 1), ("Δ Cost", 2),
                 ("Δ QALYs", 3), ("Δ LYs", 4)]:
        ws.cell(r, c, h).font = _HEADER_FONT
    for strategy in model.strategy_names:
        r += 1
        er = excel_rows[strategy]
        pr = py_rows[strategy]
        ws.cell(r, 1, model.strategy_labels[strategy])
        ws.cell(r, 2, f"=B{er}-B{pr}")
        ws.cell(r, 2).number_format = '0.000000'
        ws.cell(r, 3, f"=C{er}-C{pr}")
        ws.cell(r, 3).number_format = '0.000000'
        ws.cell(r, 4, f"=D{er}-D{pr}")
        ws.cell(r, 4).number_format = '0.000000'

    # Column widths
    ws.column_dimensions['A'].width = 22
    for c in 'BCDE':
        ws.column_dimensions[c].width = 16


# =====================================================================
# Helpers
# =====================================================================

def _validate_markov_excel_support(model, params) -> None:
    """Fail before simulation when model logic cannot be represented in Excel."""
    if model._custom_costs:
        raise NotImplementedError(
            "Formula-based Excel export cannot translate custom cost "
            "callbacks into auditable Excel formulas."
        )
    _require_constant_state_values(model, params)
    for strategy in model.strategy_names:
        for transition_cost in model._transition_costs:
            schedule = model._get_tc_schedule(
                transition_cost, strategy, params,
            )
            if schedule is None:
                raise NotImplementedError(
                    "Formula-based Excel export cannot translate callable "
                    f"transition cost {transition_cost['category']!r} for "
                    f"strategy {strategy!r}. No workbook was written."
                )
            if not np.all(np.isfinite(schedule)):
                raise ValueError(
                    f"Transition cost {transition_cost['category']!r} for "
                    f"strategy {strategy!r} contains non-finite values."
                )


def _validate_psm_excel_support(model, params) -> None:
    """Fail before simulation when PSM logic cannot be represented in Excel."""
    if model._custom_costs:
        raise NotImplementedError(
            "Formula-based Excel export cannot translate custom cost "
            "callbacks into auditable Excel formulas."
        )
    _require_constant_state_values(model, params)


def _write_survival_curve_inputs(ws, row, endpoint, curve):
    """Write supported curve parameters and return a formula specification."""
    from ..survival import (
        AcceleratedFailureTime,
        Exponential,
        GeneralizedGamma,
        Gompertz,
        KaplanMeier,
        LogLogistic,
        PiecewiseExponential,
        ProportionalHazards,
        SurvLogNormal,
        Weibull,
    )

    def write_param(label, value):
        nonlocal row
        ws.cell(row, 1, label)
        cell = ws.cell(row, 2, value)
        cell.fill = _INPUT_FILL
        ref = f"$B${row}"
        row += 1
        return ref

    prefix = endpoint
    if isinstance(curve, ProportionalHazards):
        hr_ref = write_param(f"{prefix} / PH hazard ratio", curve.hr)
        base, row = _write_survival_curve_inputs(
            ws, row, f"{prefix} / baseline", curve.baseline,
        )
        if base is None:
            return None, row
        return {"type": "ph", "baseline": base, "hr": hr_ref}, row
    if isinstance(curve, AcceleratedFailureTime):
        af_ref = write_param(f"{prefix} / AFT acceleration factor", curve.af)
        base, row = _write_survival_curve_inputs(
            ws, row, f"{prefix} / baseline", curve.baseline,
        )
        if base is None:
            return None, row
        return {"type": "aft", "baseline": base, "af": af_ref}, row
    if isinstance(curve, Exponential):
        return {
            "type": "exponential",
            "rate": write_param(f"{prefix} / Exponential rate", curve.rate),
        }, row
    if isinstance(curve, Weibull):
        return {
            "type": "weibull",
            "shape": write_param(f"{prefix} / Weibull shape", curve.shape),
            "scale": write_param(f"{prefix} / Weibull scale", curve.scale),
        }, row
    if isinstance(curve, LogLogistic):
        return {
            "type": "loglogistic",
            "shape": write_param(f"{prefix} / Log-logistic shape", curve.shape),
            "scale": write_param(f"{prefix} / Log-logistic scale", curve.scale),
        }, row
    if isinstance(curve, SurvLogNormal):
        return {
            "type": "lognormal",
            "meanlog": write_param(f"{prefix} / Log-normal meanlog", curve.meanlog),
            "sdlog": write_param(f"{prefix} / Log-normal sdlog", curve.sdlog),
        }, row
    if isinstance(curve, Gompertz):
        return {
            "type": "gompertz",
            "shape": write_param(f"{prefix} / Gompertz shape", curve.shape),
            "rate": write_param(f"{prefix} / Gompertz rate", curve.rate),
        }, row
    if isinstance(curve, GeneralizedGamma):
        return {
            "type": "generalized_gamma",
            "mu": write_param(f"{prefix} / Generalized gamma mu", curve.mu),
            "sigma": write_param(
                f"{prefix} / Generalized gamma sigma", curve.sigma,
            ),
            "q": write_param(f"{prefix} / Generalized gamma Q", curve.Q),
        }, row
    if isinstance(curve, PiecewiseExponential):
        breakpoints = [
            write_param(f"{prefix} / Breakpoint {i + 1}", float(value))
            for i, value in enumerate(curve.breakpoints)
        ]
        rates = [
            write_param(f"{prefix} / Rate {i + 1}", float(value))
            for i, value in enumerate(curve.rates)
        ]
        return {
            "type": "piecewise_exponential",
            "breakpoints": breakpoints,
            "rates": rates,
        }, row
    if isinstance(curve, KaplanMeier):
        ws.cell(row, 1, f"{prefix} / Kaplan-Meier data")
        ws.cell(row, 2, "Time").font = _HEADER_FONT
        ws.cell(row, 3, "Survival").font = _HEADER_FONT
        row += 1
        first_data_row = row
        for time, survival in zip(curve.times, curve.surv):
            for column, value in ((2, time), (3, survival)):
                cell = ws.cell(row, column, float(value))
                cell.fill = _INPUT_FILL
                cell.number_format = _FMT_PROB
            row += 1
        last_data_row = row - 1
        tail_rate = None
        if curve.extrapolation == "exponential":
            tail_rate = write_param(
                f"{prefix} / Exponential tail rate", curve._tail_rate,
            )
        return {
            "type": "kaplan_meier",
            "times": f"$B${first_data_row}:$B${last_data_row}",
            "survival": f"$C${first_data_row}:$C${last_data_row}",
            "last_time": f"$B${last_data_row}",
            "last_survival": f"$C${last_data_row}",
            "extrapolation": curve.extrapolation,
            "tail_rate": tail_rate,
        }, row

    ws.cell(row, 1, f"{prefix} / {type(curve).__name__}")
    ws.cell(row, 2, "External survival inputs below").font = _NOTE_FONT
    return None, row + 1


def _survival_formula(spec, time_ref):
    """Translate a supported survival specification into one Excel formula."""
    kind = spec["type"]
    if kind == "ph":
        base = _survival_formula(spec["baseline"], time_ref)[1:]
        return f"=({base})^{spec['hr']}"
    if kind == "aft":
        return _survival_formula(
            spec["baseline"], f"({time_ref}/{spec['af']})",
        )
    if kind == "exponential":
        return f"=EXP(-{spec['rate']}*{time_ref})"
    if kind == "weibull":
        return f"=EXP(-({time_ref}/{spec['scale']})^{spec['shape']})"
    if kind == "loglogistic":
        return f"=1/(1+({time_ref}/{spec['scale']})^{spec['shape']})"
    if kind == "lognormal":
        return (
            f"=IF({time_ref}=0,1,1-_xlfn.NORM.S.DIST("
            f"(LN({time_ref})-{spec['meanlog']})/{spec['sdlog']},TRUE))"
        )
    if kind == "gompertz":
        return (
            f"=IF(ABS({spec['shape']})<1E-12,"
            f"EXP(-{spec['rate']}*{time_ref}),"
            f"EXP(-{spec['rate']}/{spec['shape']}*"
            f"(EXP({spec['shape']}*{time_ref})-1)))"
        )
    if kind == "generalized_gamma":
        q = spec["q"]
        mu = spec["mu"]
        sigma = spec["sigma"]
        gamma_scale = f"EXP({mu}+{sigma}*LN({q}^2)/{q})"
        u = f"({time_ref}/({gamma_scale}))^({q}/{sigma})"
        gamma_cdf = f"_xlfn.GAMMA.DIST({u},1/({q}^2),1,TRUE)"
        lognormal = (
            f"1-_xlfn.NORM.S.DIST((LN({time_ref})-{mu})/{sigma},TRUE)"
        )
        return (
            f"=IF({time_ref}=0,1,IF(ABS({q})<1E-10,{lognormal},"
            f"IF({q}>0,1-{gamma_cdf},{gamma_cdf})))"
        )
    if kind == "piecewise_exponential":
        terms = []
        previous = "0"
        for index, rate in enumerate(spec["rates"]):
            if index < len(spec["breakpoints"]):
                breakpoint = spec["breakpoints"][index]
                duration = f"MAX(0,MIN({time_ref},{breakpoint})-{previous})"
                previous = breakpoint
            else:
                duration = f"MAX(0,{time_ref}-{previous})"
            terms.append(f"{rate}*{duration}")
        return f"=EXP(-({'+'.join(terms)}))"
    if kind == "kaplan_meier":
        if spec["extrapolation"] == "exponential":
            beyond = f"EXP(-{spec['tail_rate']}*{time_ref})"
        else:
            beyond = spec["last_survival"]
        return (
            f"=IF({time_ref}>{spec['last_time']},{beyond},"
            f"LOOKUP({time_ref},{spec['times']},{spec['survival']}))"
        )
    raise ValueError(f"Unsupported survival formula specification: {kind!r}")


def _require_constant_state_values(model, params) -> None:
    """Reject callbacks that cannot be represented by one Excel input vector."""
    for strategy in model.strategy_names:
        for category, definition in model._costs.items():
            first = model._resolve_state_values(
                definition.values, strategy, params, 0
            )
            for interval in range(1, model.n_cycles):
                current = model._resolve_state_values(
                    definition.values, strategy, params, interval
                )
                if not np.allclose(first, current, atol=1e-12, rtol=0):
                    raise NotImplementedError(
                        f"Formula-based Excel export cannot yet translate "
                        f"time-varying state cost {category!r} for strategy "
                        f"{strategy!r}. No workbook was written."
                    )

        first_utility = model._get_utilities(strategy, params, 0)
        for interval in range(1, model.n_cycles):
            current_utility = model._get_utilities(strategy, params, interval)
            if not np.allclose(first_utility, current_utility, atol=1e-12, rtol=0):
                raise NotImplementedError(
                    f"Formula-based Excel export cannot yet translate "
                    f"time-varying utilities for strategy {strategy!r}. "
                    "No workbook was written."
                )


def _discount_formula(time_expression: str, rate_ref: str, convention: str) -> str:
    """Build one transparent Excel discount-factor formula."""
    if convention == "continuous":
        return f"=EXP(-{rate_ref}*{time_expression})"
    return f"=1/(1+{rate_ref})^{time_expression}"


def _enable_excel_recalculation(workbook) -> None:
    """Ask Excel to recalculate every formula when the workbook opens."""
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"


def _write_setting(ws, row, label, value) -> int:
    """Write a setting row with input styling. Returns the row number."""
    fixed = label in {"Discount Convention", "N Cycles", "Half-cycle Correction", "Initial State"}
    ws.cell(row, 1, label + (" [fixed; regenerate workbook]" if fixed else ""))
    c = ws.cell(row, 2, value)
    if fixed:
        c.font = _NOTE_FONT
    else:
        c.fill = _INPUT_FILL
    return row


def _safe_sheet(name: str) -> str:
    """Truncate sheet name to 31 chars (Excel limit)."""
    return name[:31]
