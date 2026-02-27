"""
Formula-based Excel model export for cross-validation.
======================================================

Creates Excel workbooks that independently compute Markov or PSM models
using **Excel formulas**, enabling cross-validation against Python results.

The Excel file contains:

- **Input section**: Parameters, transition matrix, costs, utilities (values)
- **Calculation section**: State trace, costs, QALYs, discounting (formulas)
- **Summary sheet**: Totals from Excel formulas, Python-computed totals,
  and a Difference row so you can instantly verify they match.

Supports:

- **MarkovModel**: Full formula-based trace (time-homogeneous) or
  Python trace + formula-based downstream (time-varying)
- **PSMModel**: Python survival values + formula-based state probabilities,
  costs, QALYs, discounting

Usage
-----
>>> result = model.run_base_case()
>>> ph.export_excel_model(result, "verification.xlsx")
# or
>>> ph.export_excel_model(model, "verification.xlsx")
"""

from __future__ import annotations

import numpy as np
import pandas as pd
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter as CL
from openpyxl.styles import Font, PatternFill

# -------------------------------------------------------------------------
# Styles
# -------------------------------------------------------------------------
_TITLE_FONT = Font(bold=True, size=14)
_SECTION_FONT = Font(bold=True, size=11, underline="single")
_HEADER_FONT = Font(bold=True)
_NOTE_FONT = Font(italic=True, color="999999")
_INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")  # light yellow
_FMT_COST = '#,##0.00'
_FMT_PROB = '0.000000'
_FMT_ICER = '#,##0'


# =====================================================================
# Public API
# =====================================================================

def export_excel_model(model_or_result, filepath: str, params: dict = None):
    """Export a formula-based Excel model for cross-validation.

    Creates an Excel workbook whose calculation cells use **Excel
    formulas** (SUMPRODUCT, SUM, IF, etc.) that independently replicate
    the Python computation. A Summary sheet shows both Excel and Python
    totals plus their difference.

    Parameters
    ----------
    model_or_result : MarkovModel, PSMModel, BaseResult, or PSMBaseResult
        The model (uses base-case params) or a result object.
    filepath : str
        Output path (.xlsx).
    params : dict, optional
        Override parameter values.  Ignored when a result is passed.

    Raises
    ------
    TypeError
        If the model type is not MarkovModel or PSMModel.

    Examples
    --------
    >>> result = model.run_base_case()
    >>> ph.export_excel_model(result, "verification.xlsx")
    """
    from .model import MarkovModel
    from .psm import PSMModel
    from .results import BaseResult, PSMBaseResult

    result = None
    if isinstance(model_or_result, (BaseResult, PSMBaseResult)):
        model = model_or_result.model
        params = model_or_result.params
        result = model_or_result
    else:
        model = model_or_result
        if params is None:
            params = {name: p.base for name, p in model.params.items()}

    # Pre-compute Python results for comparison
    py_results = model._simulate_single(params)

    if isinstance(model, MarkovModel):
        _build_markov_excel(model, filepath, params, py_results)
    elif isinstance(model, PSMModel):
        _build_psm_excel(model, filepath, params, py_results)
    else:
        raise TypeError(
            f"Excel model export supports MarkovModel and PSMModel, "
            f"got {type(model).__name__}. "
            f"DES and MicroSim are individual-based and cannot be "
            f"replicated with Excel formulas."
        )


# =====================================================================
# Markov Model
# =====================================================================

def _build_markov_excel(model, filepath, params, py_results):
    """Build formula-based Excel workbook for a Markov model."""
    wb = Workbook()

    n = model.n_states
    states = model.states
    n_cycles = model.n_cycles
    cl_val = model.cycle_length
    dr_c = model.dr_costs
    dr_q = model.dr_qalys
    hcc_flag = model.half_cycle_correction
    initial_idx = model.initial_state_idx

    cost_cats = list(model._costs.keys())
    n_cats = len(cost_cats)

    # --- Detect time-homogeneity ---
    is_time_homo = _check_time_homo_trans(model, params, n_cycles)

    # --- Per-strategy sheets ---
    summary_refs = {}

    for s_idx, strategy in enumerate(model.strategy_names):
        label = model.strategy_labels[strategy]
        sname = _safe_sheet(f"Calc_{label}")

        if s_idx == 0:
            ws = wb.active
            ws.title = sname
        else:
            ws = wb.create_sheet(title=sname)

        # Evaluate transition matrix (cycle 1 for time-homo)
        P = model._get_transition_matrix(strategy, params, 1)
        PT = P.T  # rows=to-state  cols=from-state

        # ============================================
        # INPUT SECTION (yellow-filled cells)
        # ============================================
        r = 1
        ws.cell(r, 1, f"PyHEOR Markov 验证模型 — {label}").font = _TITLE_FONT
        r += 2

        # -- Settings --
        ws.cell(r, 1, "模型设置").font = _SECTION_FONT
        r += 1
        ROW_DR_C = _write_setting(ws, r, "Discount Rate (Costs)", dr_c); r += 1
        ROW_DR_Q = _write_setting(ws, r, "Discount Rate (QALYs)", dr_q); r += 1
        ROW_CL   = _write_setting(ws, r, "Cycle Length (years)", cl_val); r += 1
        _write_setting(ws, r, "N Cycles", n_cycles); r += 1
        _write_setting(ws, r, "Half-cycle Correction", "Yes" if hcc_flag else "No"); r += 1
        _write_setting(ws, r, "Initial State", states[initial_idx]); r += 2

        # -- Transition Matrix (Transposed: rows=to, cols=from) --
        ws.cell(r, 1, "转移概率矩阵 (P^T: 行=目标态, 列=来源态)").font = _SECTION_FONT
        if not is_time_homo:
            ws.cell(r, n + 3,
                    "⚠ 时变模型: 仅显示第1周期矩阵，Trace为Python预计算值"
                    ).font = Font(italic=True, color="FF0000")
        r += 1
        ws.cell(r, 1, "To \\ From")
        for j in range(n):
            ws.cell(r, 2 + j, states[j]).font = _HEADER_FONT
        r += 1

        MATRIX_R0 = r  # first data row of matrix
        for i in range(n):
            ws.cell(r, 1, states[i])
            for j in range(n):
                c = ws.cell(r, 2 + j, PT[i, j])
                c.fill = _INPUT_FILL
                c.number_format = _FMT_PROB
            r += 1
        r += 1

        # -- State Costs --
        ws.cell(r, 1, "状态费用 (年度费率)").font = _SECTION_FONT
        r += 1
        for j in range(n):
            ws.cell(r, 2 + j, states[j]).font = _HEADER_FONT
        r += 1

        cost_input_rows = {}
        for cat in cost_cats:
            vec = model._get_state_costs(cat, strategy, params, 0)
            ws.cell(r, 1, cat)
            for j in range(n):
                c = ws.cell(r, 2 + j, vec[j])
                c.fill = _INPUT_FILL
                c.number_format = _FMT_COST
            cost_input_rows[cat] = r
            r += 1
        if not cost_cats:
            ws.cell(r, 1, "(无)")
            r += 1
        r += 1

        # -- Utility Weights --
        ws.cell(r, 1, "效用权重").font = _SECTION_FONT
        r += 1
        for j in range(n):
            ws.cell(r, 2 + j, states[j]).font = _HEADER_FONT
        r += 1
        u_vec = model._get_utilities(strategy, params, 0)
        ws.cell(r, 1, "Utility")
        for j in range(n):
            c = ws.cell(r, 2 + j, u_vec[j])
            c.fill = _INPUT_FILL
        UTIL_ROW = r
        r += 1

        # -- Alive mask --
        ws.cell(r, 1, "Alive")
        alive = [1.0 if i in model._alive_states else 0.0 for i in range(n)]
        for j in range(n):
            c = ws.cell(r, 2 + j, alive[j])
            c.fill = _INPUT_FILL
        ALIVE_ROW = r
        r += 2

        # ============================================
        # CALCULATION TABLE (green-ish formulas)
        # ============================================
        ws.cell(r, 1, "计算区 (以下均为 Excel 公式)").font = _SECTION_FONT
        r += 1

        HDR = r  # header row
        D0 = HDR + 1  # first data row (cycle 0)

        # -- Column mapping --
        COL_CYC = 1
        COL_TIM = 2
        tc = 3                                     # trace col start
        te = tc + n - 1                            # trace col end
        COL_RS = te + 1                            # row sum
        COL_DFC = COL_RS + 1                       # DF cost
        COL_DFQ = COL_DFC + 1                      # DF qaly
        COL_HCC = COL_DFQ + 1

        craw = {}; cdisc = {}
        cc = COL_HCC + 1
        for cat in cost_cats:
            craw[cat] = cc; cdisc[cat] = cc + 1
            cc += 2

        COL_TC = cc;  cc += 1         # total discounted cost
        COL_QR = cc;  cc += 1         # QALY raw
        COL_QD = cc;  cc += 1         # QALY discounted
        COL_LR = cc;  cc += 1         # LY raw
        COL_LD = cc                   # LY discounted

        # -- Headers --
        hdrs = [
            (COL_CYC, "Cycle"), (COL_TIM, "Time(yr)"),
        ]
        for j in range(n):
            hdrs.append((tc + j, f"P({states[j]})"))
        hdrs += [
            (COL_RS, "RowSum"), (COL_DFC, "DF(cost)"),
            (COL_DFQ, "DF(qaly)"), (COL_HCC, "HCC"),
        ]
        for cat in cost_cats:
            hdrs += [(craw[cat], f"{cat}(raw)"), (cdisc[cat], f"{cat}(disc)")]
        hdrs += [
            (COL_TC, "TotalCost(disc)"),
            (COL_QR, "QALY(raw)"), (COL_QD, "QALY(disc)"),
            (COL_LR, "LY(raw)"), (COL_LD, "LY(disc)"),
        ]
        for col, txt in hdrs:
            ws.cell(HDR, col, txt).font = _HEADER_FONT

        # Refs that stay constant across rows
        drc = f"$B${ROW_DR_C}"
        drq = f"$B${ROW_DR_Q}"
        cl_ref = f"$B${ROW_CL}"
        # cost vector refs  {cat: "$B$rr:${CL(1+n)}$rr"}
        cvr = {cat: f"${CL(2)}${cost_input_rows[cat]}:${CL(1+n)}${cost_input_rows[cat]}"
               for cat in cost_cats}
        util_r = f"${CL(2)}${UTIL_ROW}:${CL(1+n)}${UTIL_ROW}"
        alive_r = f"${CL(2)}${ALIVE_ROW}:${CL(1+n)}${ALIVE_ROW}"

        # P^T row ref for each "to state j"  (time-homo only)
        pt_rows = {}
        for j in range(n):
            mr = MATRIX_R0 + j
            pt_rows[j] = f"${CL(2)}${mr}:${CL(1+n)}${mr}"

        # -- Data Rows --
        py_trace = py_results[strategy]['trace']  # for time-varying fallback

        for t in range(n_cycles + 1):
            rr = D0 + t  # current row

            # Cycle & Time
            ws.cell(rr, COL_CYC, t)
            ws.cell(rr, COL_TIM, f"={CL(COL_CYC)}{rr}*{cl_ref}")

            # Trace ---
            tr_range = f"{CL(tc)}{rr}:{CL(te)}{rr}"  # same-row trace (for costs)
            if t == 0:
                for j in range(n):
                    ws.cell(rr, tc + j, 1.0 if j == initial_idx else 0.0)
            else:
                if is_time_homo:
                    prev_tr = f"{CL(tc)}{rr - 1}:{CL(te)}{rr - 1}"
                    for j in range(n):
                        ws.cell(rr, tc + j,
                                f"=SUMPRODUCT({prev_tr},{pt_rows[j]})")
                else:
                    # Time-varying: write Python trace values
                    for j in range(n):
                        ws.cell(rr, tc + j, float(py_trace[t, j]))

            # Row Sum
            ws.cell(rr, COL_RS, f"=SUM({CL(tc)}{rr}:{CL(te)}{rr})")

            # Discount Factors
            time_c = f"{CL(COL_TIM)}{rr}"
            ws.cell(rr, COL_DFC, f"=1/(1+{drc})^{time_c}")
            ws.cell(rr, COL_DFQ, f"=1/(1+{drq})^{time_c}")

            # HCC
            if hcc_flag and (t == 0 or t == n_cycles):
                ws.cell(rr, COL_HCC, 0.5)
            else:
                ws.cell(rr, COL_HCC, 1.0)

            df_c_cell = f"{CL(COL_DFC)}{rr}"
            df_q_cell = f"{CL(COL_DFQ)}{rr}"
            hcc_cell = f"{CL(COL_HCC)}{rr}"

            # --- Costs ---
            for cat in cost_cats:
                cdef = model._costs[cat]
                base = f"SUMPRODUCT({tr_range},{cvr[cat]})"

                if cdef.method == "wlos":
                    inner = f"{base}*{cl_ref}"
                else:
                    inner = base

                if cdef.first_cycle_only:
                    raw_f = f"=IF({CL(COL_CYC)}{rr}=0,{inner},0)"
                elif cdef.apply_cycles is not None:
                    # Specific cycles: store value
                    val = py_results[strategy]['costs_by_cycle'].get(
                        cat, np.zeros(n_cycles + 1))[t]
                    ws.cell(rr, craw[cat], val)
                    ws.cell(rr, craw[cat]).number_format = _FMT_COST
                    raw_ref = f"{CL(craw[cat])}{rr}"
                    ws.cell(rr, cdisc[cat],
                            f"={raw_ref}*{df_c_cell}*{hcc_cell}")
                    ws.cell(rr, cdisc[cat]).number_format = _FMT_COST
                    continue
                else:
                    raw_f = f"={inner}"

                ws.cell(rr, craw[cat], raw_f)
                ws.cell(rr, craw[cat]).number_format = _FMT_COST

                raw_ref = f"{CL(craw[cat])}{rr}"
                if cdef.method == "starting":
                    ws.cell(rr, cdisc[cat], f"={raw_ref}*{df_c_cell}")
                else:
                    ws.cell(rr, cdisc[cat],
                            f"={raw_ref}*{df_c_cell}*{hcc_cell}")
                ws.cell(rr, cdisc[cat]).number_format = _FMT_COST

            # Total discounted cost
            if n_cats:
                expr = "+".join(f"{CL(cdisc[c])}{rr}" for c in cost_cats)
                ws.cell(rr, COL_TC, f"={expr}")
            else:
                ws.cell(rr, COL_TC, 0)
            ws.cell(rr, COL_TC).number_format = _FMT_COST

            # --- QALYs ---
            ws.cell(rr, COL_QR,
                    f"=SUMPRODUCT({tr_range},{util_r})*{cl_ref}")
            ws.cell(rr, COL_QD,
                    f"={CL(COL_QR)}{rr}*{df_q_cell}*{hcc_cell}")
            ws.cell(rr, COL_QD).number_format = _FMT_PROB

            # --- LYs ---
            ws.cell(rr, COL_LR,
                    f"=SUMPRODUCT({tr_range},{alive_r})*{cl_ref}")
            ws.cell(rr, COL_LD,
                    f"={CL(COL_LR)}{rr}*{df_q_cell}*{hcc_cell}")

        # -- Handle transition costs (add as pre-computed values) --
        tc_note_row = None
        if model._transition_costs:
            # Add extra cost columns for transition costs
            tc_col_start = COL_LD + 2
            tc_cats_done = set()
            tc_col_map = {}
            for tc_def in model._transition_costs:
                tcat = tc_def['category']
                if tcat not in tc_cats_done:
                    ws.cell(HDR, tc_col_start, f"{tcat}(tc,disc)")
                    ws.cell(HDR, tc_col_start).font = _HEADER_FONT
                    tc_col_map[tcat] = tc_col_start
                    tc_col_start += 1
                    tc_cats_done.add(tcat)

            # Write pre-computed discounted transition costs
            for tcat in tc_col_map:
                if tcat in py_results[strategy].get('discounted_costs', {}):
                    arr = py_results[strategy]['discounted_costs'][tcat]
                    for t in range(n_cycles + 1):
                        rr = D0 + t
                        ws.cell(rr, tc_col_map[tcat], float(arr[t]))
                        ws.cell(rr, tc_col_map[tcat]).number_format = _FMT_COST

            tc_note_row = D0 + n_cycles + 2

        # -- Totals Row --
        TR = D0 + n_cycles + 1
        DL = D0 + n_cycles  # data last row
        ws.cell(TR, COL_CYC, "TOTAL").font = _HEADER_FONT

        sum_cols = (
            [cdisc[c] for c in cost_cats]
            + [COL_TC, COL_QD, COL_LD]
        )
        for sc in sum_cols:
            ws.cell(TR, sc,
                    f"=SUM({CL(sc)}{D0}:{CL(sc)}{DL})")
            ws.cell(TR, sc).font = _HEADER_FONT
            ws.cell(TR, sc).number_format = _FMT_COST

        # If transition costs exist, add them to total cost
        if model._transition_costs and tc_col_map:
            # Adjust total cost in totals row to include tc
            tc_sum_parts = []
            for tcat, tcc in tc_col_map.items():
                ws.cell(TR, tcc, f"=SUM({CL(tcc)}{D0}:{CL(tcc)}{DL})")
                ws.cell(TR, tcc).font = _HEADER_FONT
                ws.cell(TR, tcc).number_format = _FMT_COST
                tc_sum_parts.append(f"{CL(tcc)}{TR}")
            # Overwrite total cost to include transition costs
            base_tc = f"SUM({CL(COL_TC)}{D0}:{CL(COL_TC)}{DL})"
            tc_extra = "+".join(tc_sum_parts)
            ws.cell(TR, COL_TC, f"={base_tc}+{tc_extra}")

        if tc_note_row:
            ws.cell(tc_note_row, 1,
                    "Note: Transition costs (tc) are pre-computed from Python "
                    "(schedule-based convolution cannot be replicated in Excel formulas)."
                    ).font = _NOTE_FONT

        # Store refs for Summary
        summary_refs[strategy] = {
            'sheet': sname,
            'cost': f"'{sname}'!{CL(COL_TC)}{TR}",
            'qaly': f"'{sname}'!{CL(COL_QD)}{TR}",
            'ly':   f"'{sname}'!{CL(COL_LD)}{TR}",
        }

        # Freeze & widths
        ws.freeze_panes = ws.cell(HDR + 1, tc)
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 10

    # ============================================
    # SUMMARY SHEET
    # ============================================
    _build_summary_sheet(wb, model, summary_refs, py_results)

    wb.save(filepath)
    print(f"✅ Excel 验证模型已导出: {filepath}")
    if not is_time_homo:
        print("   ⚠ 时变模型: Trace 为 Python 预计算值; "
              "费用/QALY/贴现/ICER 仍为 Excel 公式。")


# =====================================================================
# PSM Model
# =====================================================================

def _build_psm_excel(model, filepath, params, py_results):
    """Build formula-based Excel workbook for a PSM model."""
    wb = Workbook()

    n = model.n_states
    states = model.states
    n_ep = model.n_endpoints
    endpoints = model.survival_endpoints
    n_cycles = model.n_cycles
    cl_val = model.cycle_length
    dr_c = model.dr_costs
    dr_q = model.dr_qalys
    hcc_flag = model.half_cycle_correction

    cost_cats = list(model._costs.keys())
    n_cats = len(cost_cats)

    summary_refs = {}

    for s_idx, strategy in enumerate(model.strategy_names):
        label = model.strategy_labels[strategy]
        sname = _safe_sheet(f"Calc_{label}")

        if s_idx == 0:
            ws = wb.active
            ws.title = sname
        else:
            ws = wb.create_sheet(title=sname)

        # ============================================
        # INPUT SECTION
        # ============================================
        r = 1
        ws.cell(r, 1, f"PyHEOR PSM 验证模型 — {label}").font = _TITLE_FONT
        r += 2

        ws.cell(r, 1, "模型设置").font = _SECTION_FONT; r += 1
        ROW_DR_C = _write_setting(ws, r, "Discount Rate (Costs)", dr_c); r += 1
        ROW_DR_Q = _write_setting(ws, r, "Discount Rate (QALYs)", dr_q); r += 1
        ROW_CL   = _write_setting(ws, r, "Cycle Length (years)", cl_val); r += 1
        _write_setting(ws, r, "N Cycles", n_cycles); r += 1
        _write_setting(ws, r, "Half-cycle Correction", "Yes" if hcc_flag else "No"); r += 2

        # State Costs
        ws.cell(r, 1, "状态费用 (年度费率)").font = _SECTION_FONT; r += 1
        for j in range(n):
            ws.cell(r, 2 + j, states[j]).font = _HEADER_FONT
        r += 1
        cost_input_rows = {}
        for cat in cost_cats:
            vec = model._get_state_costs(cat, strategy, params, 0)
            ws.cell(r, 1, cat)
            for j in range(n):
                c = ws.cell(r, 2 + j, vec[j])
                c.fill = _INPUT_FILL
                c.number_format = _FMT_COST
            cost_input_rows[cat] = r
            r += 1
        if not cost_cats:
            ws.cell(r, 1, "(无)"); r += 1
        r += 1

        # Utility
        ws.cell(r, 1, "效用权重").font = _SECTION_FONT; r += 1
        for j in range(n):
            ws.cell(r, 2 + j, states[j]).font = _HEADER_FONT
        r += 1
        u_vec = model._get_utilities(strategy, params, 0)
        ws.cell(r, 1, "Utility")
        for j in range(n):
            c = ws.cell(r, 2 + j, u_vec[j])
            c.fill = _INPUT_FILL
        UTIL_ROW = r; r += 1

        ws.cell(r, 1, "Alive")
        alive = [1.0 if i in model._alive_states else 0.0 for i in range(n)]
        for j in range(n):
            c = ws.cell(r, 2 + j, alive[j])
            c.fill = _INPUT_FILL
        ALIVE_ROW = r; r += 2

        # ============================================
        # CALCULATION TABLE
        # ============================================
        ws.cell(r, 1, "计算区 (生存曲线为Python值, 状态概率/费用/QALY为Excel公式)").font = _SECTION_FONT
        r += 1

        HDR = r
        D0 = HDR + 1

        # Column layout
        COL_CYC = 1
        COL_TIM = 2
        # Survival curves (values)
        surv_start = 3
        surv_end = surv_start + n_ep - 1
        # State probabilities (formulas)
        sp_start = surv_end + 1
        sp_end = sp_start + n - 1
        COL_RS = sp_end + 1
        COL_DFC = COL_RS + 1
        COL_DFQ = COL_DFC + 1
        COL_HCC = COL_DFQ + 1

        craw = {}; cdisc = {}
        cc = COL_HCC + 1
        for cat in cost_cats:
            craw[cat] = cc; cdisc[cat] = cc + 1; cc += 2

        COL_TC = cc; cc += 1
        COL_QR = cc; cc += 1
        COL_QD = cc; cc += 1
        COL_LR = cc; cc += 1
        COL_LD = cc

        # Headers
        headers = [(COL_CYC, "Cycle"), (COL_TIM, "Time(yr)")]
        for j, ep in enumerate(endpoints):
            headers.append((surv_start + j, f"S({ep})"))
        for j, st in enumerate(states):
            headers.append((sp_start + j, f"P({st})"))
        headers += [
            (COL_RS, "RowSum"), (COL_DFC, "DF(cost)"),
            (COL_DFQ, "DF(qaly)"), (COL_HCC, "HCC"),
        ]
        for cat in cost_cats:
            headers += [(craw[cat], f"{cat}(raw)"), (cdisc[cat], f"{cat}(disc)")]
        headers += [
            (COL_TC, "TotalCost(disc)"),
            (COL_QR, "QALY(raw)"), (COL_QD, "QALY(disc)"),
            (COL_LR, "LY(raw)"), (COL_LD, "LY(disc)"),
        ]
        for col, txt in headers:
            ws.cell(HDR, col, txt).font = _HEADER_FONT

        # Constant refs
        drc = f"$B${ROW_DR_C}"
        drq = f"$B${ROW_DR_Q}"
        cl_ref = f"$B${ROW_CL}"
        cvr = {cat: f"${CL(2)}${cost_input_rows[cat]}:${CL(1+n)}${cost_input_rows[cat]}"
               for cat in cost_cats}
        util_r = f"${CL(2)}${UTIL_ROW}:${CL(1+n)}${UTIL_ROW}"
        alive_r = f"${CL(2)}${ALIVE_ROW}:${CL(1+n)}${ALIVE_ROW}"

        # Get survival values from Python
        surv_data = py_results[strategy]['survival_curves']  # {ep: array}

        # -- Data Rows --
        for t in range(n_cycles + 1):
            rr = D0 + t

            ws.cell(rr, COL_CYC, t)
            ws.cell(rr, COL_TIM, f"={CL(COL_CYC)}{rr}*{cl_ref}")

            # Survival curve VALUES
            for j, ep in enumerate(endpoints):
                ws.cell(rr, surv_start + j, float(surv_data[ep][t]))
                ws.cell(rr, surv_start + j).number_format = _FMT_PROB

            # State probability FORMULAS
            # state[0] = S(endpoint_0)
            ws.cell(rr, sp_start,
                    f"={CL(surv_start)}{rr}")

            # state[k] = MAX(S(endpoint_k) - S(endpoint_{k-1}), 0) for k=1..n_ep-1
            for k in range(1, n_ep):
                ws.cell(rr, sp_start + k,
                        f"=MAX({CL(surv_start + k)}{rr}-{CL(surv_start + k - 1)}{rr},0)")

            # state[-1] = 1 - S(last_endpoint)
            ws.cell(rr, sp_start + n - 1,
                    f"=1-{CL(surv_end)}{rr}")

            # Row Sum
            ws.cell(rr, COL_RS,
                    f"=SUM({CL(sp_start)}{rr}:{CL(sp_end)}{rr})")

            # Discount factors
            time_c = f"{CL(COL_TIM)}{rr}"
            ws.cell(rr, COL_DFC, f"=1/(1+{drc})^{time_c}")
            ws.cell(rr, COL_DFQ, f"=1/(1+{drq})^{time_c}")

            # HCC
            if hcc_flag and (t == 0 or t == n_cycles):
                ws.cell(rr, COL_HCC, 0.5)
            else:
                ws.cell(rr, COL_HCC, 1.0)

            df_c_cell = f"{CL(COL_DFC)}{rr}"
            df_q_cell = f"{CL(COL_DFQ)}{rr}"
            hcc_cell = f"{CL(COL_HCC)}{rr}"
            sp_range = f"{CL(sp_start)}{rr}:{CL(sp_end)}{rr}"

            # Costs
            for cat in cost_cats:
                cdef = model._costs[cat]
                base = f"SUMPRODUCT({sp_range},{cvr[cat]})"

                if cdef.method == "wlos":
                    inner = f"{base}*{cl_ref}"
                else:
                    inner = base

                if cdef.first_cycle_only:
                    raw_f = f"=IF({CL(COL_CYC)}{rr}=0,{inner},0)"
                elif cdef.apply_cycles is not None:
                    val = py_results[strategy]['costs_by_cycle'].get(
                        cat, np.zeros(n_cycles + 1))[t]
                    ws.cell(rr, craw[cat], val)
                    ws.cell(rr, craw[cat]).number_format = _FMT_COST
                    raw_ref = f"{CL(craw[cat])}{rr}"
                    ws.cell(rr, cdisc[cat],
                            f"={raw_ref}*{df_c_cell}*{hcc_cell}")
                    ws.cell(rr, cdisc[cat]).number_format = _FMT_COST
                    continue
                else:
                    raw_f = f"={inner}"

                ws.cell(rr, craw[cat], raw_f)
                ws.cell(rr, craw[cat]).number_format = _FMT_COST
                raw_ref = f"{CL(craw[cat])}{rr}"
                if cdef.method == "starting":
                    ws.cell(rr, cdisc[cat], f"={raw_ref}*{df_c_cell}")
                else:
                    ws.cell(rr, cdisc[cat],
                            f"={raw_ref}*{df_c_cell}*{hcc_cell}")
                ws.cell(rr, cdisc[cat]).number_format = _FMT_COST

            # Total discounted cost
            if n_cats:
                expr = "+".join(f"{CL(cdisc[c])}{rr}" for c in cost_cats)
                ws.cell(rr, COL_TC, f"={expr}")
            else:
                ws.cell(rr, COL_TC, 0)
            ws.cell(rr, COL_TC).number_format = _FMT_COST

            # QALYs
            ws.cell(rr, COL_QR,
                    f"=SUMPRODUCT({sp_range},{util_r})*{cl_ref}")
            ws.cell(rr, COL_QD,
                    f"={CL(COL_QR)}{rr}*{df_q_cell}*{hcc_cell}")
            ws.cell(rr, COL_QD).number_format = _FMT_PROB

            # LYs
            ws.cell(rr, COL_LR,
                    f"=SUMPRODUCT({sp_range},{alive_r})*{cl_ref}")
            ws.cell(rr, COL_LD,
                    f"={CL(COL_LR)}{rr}*{df_q_cell}*{hcc_cell}")

        # Totals
        TR = D0 + n_cycles + 1
        DL = D0 + n_cycles
        ws.cell(TR, COL_CYC, "TOTAL").font = _HEADER_FONT
        for sc in ([cdisc[c] for c in cost_cats]
                   + [COL_TC, COL_QD, COL_LD]):
            ws.cell(TR, sc,
                    f"=SUM({CL(sc)}{D0}:{CL(sc)}{DL})")
            ws.cell(TR, sc).font = _HEADER_FONT
            ws.cell(TR, sc).number_format = _FMT_COST

        summary_refs[strategy] = {
            'sheet': sname,
            'cost': f"'{sname}'!{CL(COL_TC)}{TR}",
            'qaly': f"'{sname}'!{CL(COL_QD)}{TR}",
            'ly':   f"'{sname}'!{CL(COL_LD)}{TR}",
        }

        ws.freeze_panes = ws.cell(HDR + 1, surv_start)
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 10

    # Summary
    _build_summary_sheet(wb, model, summary_refs, py_results)

    wb.save(filepath)
    print(f"✅ Excel 验证模型已导出: {filepath}")
    print("   ℹ 生存曲线为 Python 预计算值; "
          "状态概率/费用/QALY/贴现/ICER 均为 Excel 公式。")


# =====================================================================
# Summary Sheet (shared by Markov & PSM)
# =====================================================================

def _build_summary_sheet(wb, model, summary_refs, py_results):
    """Build a Summary & Cross-Validation sheet."""
    ws = wb.create_sheet("Summary", 0)  # first sheet
    ws.cell(1, 1, "Summary & Cross-Validation").font = _TITLE_FONT

    r = 3
    # --- Excel formula results ---
    ws.cell(r, 1, "Excel 公式计算结果").font = _SECTION_FONT; r += 1
    for h, c in [("Strategy", 1), ("Total Cost", 2),
                 ("QALYs", 3), ("LYs", 4)]:
        ws.cell(r, c, h).font = _HEADER_FONT
    excel_rows = {}
    for s_idx, strategy in enumerate(model.strategy_names):
        rr = r + 1 + s_idx
        ws.cell(rr, 1, model.strategy_labels[strategy])
        ws.cell(rr, 2, f"={summary_refs[strategy]['cost']}")
        ws.cell(rr, 2).number_format = _FMT_COST
        ws.cell(rr, 3, f"={summary_refs[strategy]['qaly']}")
        ws.cell(rr, 3).number_format = _FMT_PROB
        ws.cell(rr, 4, f"={summary_refs[strategy]['ly']}")
        ws.cell(rr, 4).number_format = _FMT_PROB
        excel_rows[strategy] = rr

    # --- ICER ---
    r = rr + 2
    ws.cell(r, 1, "ICER (Excel 公式)").font = _SECTION_FONT; r += 1
    for h, c in [("Strategy", 1), ("vs", 2), ("Inc. Cost", 3),
                 ("Inc. QALYs", 4), ("ICER ($/QALY)", 5)]:
        ws.cell(r, c, h).font = _HEADER_FONT

    comp = model.strategy_names[0]
    comp_row = excel_rows[comp]
    for strategy in model.strategy_names[1:]:
        r += 1
        ir = excel_rows[strategy]
        ws.cell(r, 1, model.strategy_labels[strategy])
        ws.cell(r, 2, model.strategy_labels[comp])
        ws.cell(r, 3, f"=B{ir}-B{comp_row}")
        ws.cell(r, 3).number_format = _FMT_COST
        ws.cell(r, 4, f"=C{ir}-C{comp_row}")
        ws.cell(r, 4).number_format = _FMT_PROB
        ws.cell(r, 5, f'=IF(ABS(D{r})<0.0001,"N/A",C{r}/D{r})')
        ws.cell(r, 5).number_format = _FMT_ICER

    # --- Python results ---
    r += 2
    ws.cell(r, 1, "Python 计算结果 (对照)").font = _SECTION_FONT; r += 1
    for h, c in [("Strategy", 1), ("Total Cost", 2),
                 ("QALYs", 3), ("LYs", 4)]:
        ws.cell(r, c, h).font = _HEADER_FONT
    py_rows = {}
    for s_idx, strategy in enumerate(model.strategy_names):
        rr = r + 1 + s_idx
        pr = py_results[strategy]
        ws.cell(rr, 1, model.strategy_labels[strategy])
        ws.cell(rr, 2, sum(pr['total_costs'].values()))
        ws.cell(rr, 2).number_format = _FMT_COST
        ws.cell(rr, 3, pr['total_qalys'])
        ws.cell(rr, 3).number_format = _FMT_PROB
        ws.cell(rr, 4, pr['total_lys'])
        ws.cell(rr, 4).number_format = _FMT_PROB
        py_rows[strategy] = rr

    # --- Difference ---
    r = rr + 2
    ws.cell(r, 1, "差异 (Excel − Python)").font = _SECTION_FONT; r += 1
    for h, c in [("Strategy", 1), ("Δ Cost", 2),
                 ("Δ QALYs", 3), ("Δ LYs", 4)]:
        ws.cell(r, c, h).font = _HEADER_FONT
    for strategy in model.strategy_names:
        r += 1
        er = excel_rows[strategy]
        pr = py_rows[strategy]
        ws.cell(r, 1, model.strategy_labels[strategy])
        ws.cell(r, 2, f"=B{er}-B{pr}")
        ws.cell(r, 2).number_format = '0.000000'
        ws.cell(r, 3, f"=C{er}-C{pr}")
        ws.cell(r, 3).number_format = '0.000000'
        ws.cell(r, 4, f"=D{er}-D{pr}")
        ws.cell(r, 4).number_format = '0.000000'

    # Column widths
    ws.column_dimensions['A'].width = 22
    for c in 'BCDE':
        ws.column_dimensions[c].width = 16


# =====================================================================
# Helpers
# =====================================================================

def _check_time_homo_trans(model, params, n_cycles) -> bool:
    """Check if transition matrices are time-homogeneous."""
    for strategy in model.strategy_names:
        P1 = model._get_transition_matrix(strategy, params, 1)
        P2 = model._get_transition_matrix(strategy, params, min(2, n_cycles))
        if not np.allclose(P1, P2, atol=1e-10):
            return False
        if n_cycles > 5:
            Pm = model._get_transition_matrix(strategy, params, n_cycles // 2)
            if not np.allclose(P1, Pm, atol=1e-10):
                return False
    return True


def _write_setting(ws, row, label, value) -> int:
    """Write a setting row with input styling. Returns the row number."""
    ws.cell(row, 1, label)
    c = ws.cell(row, 2, value)
    c.fill = _INPUT_FILL
    return row


def _safe_sheet(name: str) -> str:
    """Truncate sheet name to 31 chars (Excel limit)."""
    return name[:31]
