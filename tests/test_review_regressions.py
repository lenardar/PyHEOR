import numpy as np
import pytest
from openpyxl import load_workbook
from pyheor import MarkovModel, PSMModel, Exponential, Beta, Gamma, export_excel_model


def make_model(kind):
    kwargs = dict(states=['Alive', 'Dead'], strategies=['S'], n_cycles=2)
    if kind == 'markov':
        model = MarkovModel(**kwargs)
        model.set_transitions('S', [[1, 0], [0, 1]])
    else:
        model = PSMModel(**kwargs, survival_endpoints=['OS'])
        model.set_survival('S', 'OS', Exponential(rate=.1))
    return model


@pytest.mark.parametrize('kind', ['markov', 'psm'])
def test_late_cost_export_preserves_rate_and_discount_override(kind, tmp_path):
    model = make_model(kind)
    model.set_state_cost('late', {'Alive': 100}, apply_cycles=[1])
    path = tmp_path / 'audit.xlsx'
    export_excel_model(model, path, params={'dr_cost': .5})
    workbook = load_workbook(path)
    sheet = workbook['Calc_S']
    row = next(row for row in sheet if row[0].value == 'late')
    assert row[1].value == 100
    assert sheet['B4'].value == .5
    assert model.dr_cost == 0
    for row in sheet:
        if isinstance(row[0].value, str) and '[fixed;' in row[0].value:
            assert row[1].fill.patternType is None
    with model._attr_param_override({'dr_cost': .5}):
        expected = sum(model.run_base_case().results['S']['total_costs'].values())
    summary = workbook['Summary']
    heading = next(c.row for row in summary for c in row if c.value == 'Python 计算结果 (对照)')
    assert summary.cell(heading + 2, 2).value == pytest.approx(expected)


@pytest.mark.parametrize('kind', ['markov', 'psm'])
@pytest.mark.parametrize('value', [np.nan, np.inf, -np.inf])
@pytest.mark.parametrize('reward', ['cost', 'utility'])
def test_nonfinite_rewards_fail(kind, value, reward):
    model = make_model(kind)
    if reward == 'cost':
        model.set_state_cost('care', {'Alive': value})
    else:
        model.set_utility({'Alive': value})
    with pytest.raises(ValueError, match='Non-finite.*strategy.*interval'):
        model.run_base_case()


@pytest.mark.parametrize('factory,mean', [(Beta, .5), (Gamma, 100)])
@pytest.mark.parametrize('sd', [0, -1, np.nan, np.inf])
def test_invalid_uncertainty_is_not_reinterpreted(factory, mean, sd):
    with pytest.raises(ValueError, match='sd must be finite and positive'):
        factory(mean=mean, sd=sd)


def test_dominated_classification_agrees_across_analyses():
    model = MarkovModel(states=['Alive', 'Dead'], strategies=['SOC', 'Bad'], n_cycles=1)
    model.set_transitions('SOC', [[1, 0], [0, 1]])
    model.set_transitions('Bad', [[0, 1], [0, 1]])
    model.add_param('cost', 100, low=80, high=120)
    model.set_state_cost('init', {'Bad': {'Alive': 'cost'}}, method='starting')
    assert model.run_base_case().icer().iloc[0]['ICER'] == 'Dominated'
    psa = model.run_psa(n_sim=2, seed=1, progress=False).icer().iloc[0]
    assert np.isnan(psa['ICER'])
    assert psa['ICER Classification'] == 'Dominated'
    owsa = model.run_owsa().summary(outcome='icer').iloc[0]
    for case in ['Low', 'High', 'Base']:
        assert np.isnan(owsa[f'ICER ({case})'])
        assert owsa[f'ICER Classification ({case})'] == 'Dominated'
