"""Tests that distinguish result exports from auditable Excel models."""

from pathlib import Path

import pytest
from openpyxl import load_workbook

from pyheor import (
    Exponential,
    GeneralizedGamma,
    KaplanMeier,
    MarkovModel,
    PSMModel,
    PiecewiseExponential,
    export_excel_model,
    export_to_excel,
)


def _markov_model(n_cycles=2, hcc=True):
    model = MarkovModel(
        states=["Alive", "Dead"], strategies=["S1"],
        n_cycles=n_cycles, half_cycle_correction=hcc,
        dr_cost=0.03, dr_qaly=0.03,
    )
    model.set_transitions("S1", [[1, 0], [0, 1]])
    model.set_state_cost("care", {"Alive": 100, "Dead": 0})
    model.set_utility({"Alive": 1, "Dead": 0})
    return model


def _formula_cells(worksheet):
    return [
        cell.value
        for row in worksheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]


def test_result_export_has_one_row_per_interval(tmp_path):
    model = _markov_model(n_cycles=2)
    path = tmp_path / "results.xlsx"
    export_to_excel(model.run_base_case(), path)

    workbook = load_workbook(path, data_only=False)
    costs = workbook["Costs_S1"]
    assert costs["A2"].value == 0
    assert costs["A3"].value == 1
    assert costs["A4"].value == "TOTAL"


def test_result_export_rejects_formula_claim(tmp_path):
    path = tmp_path / "not-a-model.xlsx"
    with pytest.raises(ValueError, match="export_excel_model"):
        export_to_excel(
            _markov_model().run_base_case(), path, include_formulas=True
        )
    assert not path.exists()


def test_markov_audit_workbook_uses_formulas_and_recalculates(tmp_path):
    path = tmp_path / "markov-model.xlsx"
    export_excel_model(_markov_model(), path)

    workbook = load_workbook(path, data_only=False)
    calculation = workbook["Calc_S1"]
    formulas = _formula_cells(calculation)

    assert workbook.calculation.calcMode == "auto"
    assert workbook.calculation.fullCalcOnLoad is True
    assert any("SUMPRODUCT" in formula for formula in formulas)
    assert any("/2" in formula and "+" in formula for formula in formulas)
    assert any("/2)" in formula and "1+" in formula for formula in formulas)
    assert any('"OK","ERROR"' in formula for formula in formulas)

    header_row = next(
        row for row in calculation.iter_rows()
        if row[0].value == "Cycle" and row[1].value == "Time(yr)"
    )
    header_number = header_row[0].row
    final_trace_row = header_number + 1 + 2
    assert calculation.cell(final_trace_row, 1).value == 2
    assert calculation.cell(final_trace_row, 3).data_type == "f"
    # The final trace point is observable but does not accrue another reward.
    qaly_column = next(
        cell.column for cell in header_row if cell.value == "QALY(raw)"
    )
    assert calculation.cell(final_trace_row, qaly_column).value is None


def test_time_varying_markov_uses_interval_matrix_inputs_and_trace_formulas(tmp_path):
    model = MarkovModel(
        states=["Alive", "Dead"], strategies=["S1"], n_cycles=2,
    )
    model.set_transitions(
        "S1", lambda params, interval: [
            [1 - interval * 0.1, interval * 0.1], [0, 1]
        ]
    )
    path = tmp_path / "time-varying.xlsx"
    export_excel_model(model, path)

    calculation = load_workbook(path, data_only=False)["Calc_S1"]
    values = [cell.value for row in calculation.iter_rows() for cell in row]
    formulas = _formula_cells(calculation)
    assert "Interval 0" in values
    assert "Interval 1" in values
    assert sum("SUMPRODUCT" in formula for formula in formulas) >= 4


def test_transition_cost_and_schedule_use_excel_formulas(tmp_path):
    model = MarkovModel(
        states=["Alive", "Dead"], strategies=["S1"], n_cycles=2,
        dr_cost=0.03,
    )
    model.set_transitions("S1", [[0.8, 0.2], [0, 1]])
    model.set_transition_cost("event", "Alive", "Dead", [100, 20])
    path = tmp_path / "transition-cost.xlsx"
    export_excel_model(model, path)

    calculation = load_workbook(path, data_only=False)["Calc_S1"]
    values = [cell.value for row in calculation.iter_rows() for cell in row]
    formulas = _formula_cells(calculation)
    assert "event: Alive → Dead" in values
    assert "Offset 0" in values
    assert "Offset 1" in values
    assert any("$" in formula and "*" in formula for formula in formulas)


def test_callable_transition_cost_is_rejected_before_execution(tmp_path):
    model = _markov_model()

    def unsupported(params, interval):
        raise AssertionError("callback must not run during failed export")

    model.set_transition_cost("event", "Alive", "Dead", unsupported)
    path = tmp_path / "callable-transition-cost.xlsx"
    with pytest.raises(NotImplementedError, match="callable transition cost"):
        export_excel_model(model, path)
    assert not path.exists()


def test_psm_parametric_survival_uses_excel_formulas(tmp_path):
    model = PSMModel(
        states=["Alive", "Dead"], survival_endpoints=["OS"],
        strategies=["S1"], n_cycles=2,
    )
    model.set_survival("S1", "OS", Exponential(rate=0.1))
    model.set_state_cost("care", {"Alive": 100, "Dead": 0})
    model.set_utility({"Alive": 1, "Dead": 0})

    path = tmp_path / "psm-model.xlsx"
    export_excel_model(model, path)
    calculation = load_workbook(path, data_only=False)["Calc_S1"]
    formulas = _formula_cells(calculation)

    section_cells = [cell.value for row in calculation.iter_rows() for cell in row]
    assert "OS / Exponential rate" in section_cells
    assert any("EXP(-" in formula for formula in formulas)
    assert any("SUMPRODUCT" in formula for formula in formulas)
    assert not any(formula.startswith("=MAX(") for formula in formulas)
    assert any('"OK","ERROR"' in formula for formula in formulas)


@pytest.mark.parametrize(
    ("curve", "formula_token"),
    [
        (GeneralizedGamma(mu=0, sigma=1, Q=1), "_xlfn.GAMMA.DIST"),
        (PiecewiseExponential(breakpoints=[1], rates=[0.1, 0.2]), "MIN("),
        (
            KaplanMeier(
                times=[0, 1, 2], survival_probs=[1, 0.9, 0.8],
            ),
            "LOOKUP(",
        ),
    ],
)
def test_psm_library_survival_types_use_excel_formulas(
    tmp_path, curve, formula_token,
):
    model = PSMModel(
        states=["Alive", "Dead"], survival_endpoints=["OS"],
        strategies=["S1"], n_cycles=2,
    )
    model.set_survival("S1", "OS", curve)
    path = tmp_path / "psm-library-curve.xlsx"
    export_excel_model(model, path)

    calculation = load_workbook(path, data_only=False)["Calc_S1"]
    formulas = _formula_cells(calculation)
    assert any(formula_token in formula for formula in formulas)
